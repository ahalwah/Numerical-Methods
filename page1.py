from tkinter import *
from tkinter import messagebox
from sympy import *
import sympy as sym
from run import execute
from run import differentiate
from run import integrate
from run import curve_fitting
from run import unconstrained
from PIL import ImageTk, Image
from tkinter import filedialog
from test import simplex
from openpyxl import load_workbook
import os


root = Tk()
root.title('Numerical Methods')
root.geometry('300x500+500+100')
root.resizable(0, 0)
entries = []

options = ['Secant Method',
           'Modified Secant',
           'Newton-Raphson Method',
           'Modified Newton-Raphson',
           'Non-Linear System of Eqs']
options2 = ["Euler's Method",
            "Heun's Method",
            'Midpoint Method',
            '4th Order Runge-Kutta',
            '5th Order Runge-Kutta']
options3 = ['Single Application Trapezoidal Rule',
            'Multiple Application Trapezoidal Rule',
            "Single Application Simpson's 1/3 Rule",
            "Multiple Application Simpson's 1/3 Rule",
            "Simpson's 3/8 Rule",
            'Trapezoidal Rule with Unequal Segments']
options4 = ['Linear Regression',
            'Polynomial Regression',
            'Multiple Linear Regression',
            'Non-Linear Regression']
options5 = ['Golden Section Search',
            'Parabolic Interpolation',
            "Newton's Method",
            'Optimal Steepest Ascent Method']


pressed = 0
pressed1 = 0
pressed2 = 0
pressed3 = 0
pressed4 = 0


def b4_press():
    global D1, D2, D3, D4, frame4
    if clicked5.get() == 'Golden Section Search':
        w5 = Tk()
        w5.title(clicked5.get())
        w5.geometry('600x250+300+200')
        w5.resizable(0, 0)
        frame4 = Frame(w5)
        frame4.pack(side=TOP, expand=False)
        Label(frame4, text=clicked5.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame4, text='f(x) = ', font=(
            "Times", "12")).grid(row=1, column=0)
        D1 = Entry(frame4, width=50)
        D1.grid(row=1, column=1, columnspan=2)
        Label(frame4, text='x lower bound ', font=(
            "Times", "12")).grid(row=2, column=0)
        D2 = Entry(frame4, width=50)
        D2.grid(row=2, column=1, columnspan=2)
        Label(frame4, text='x upper bound ', font=(
            "Times", "12")).grid(row=3, column=0)
        D3 = Entry(frame4, width=50)
        D3.grid(row=3, column=1, columnspan=2)
        Button(frame4, text='calculate max', width=12,
               command=maxim_gold).grid(row=4, column=1)
        Button(frame4, text='calculate min', width=12,
               command=minim_gold).grid(row=4, column=2)
    if clicked5.get() == 'Parabolic Interpolation':
        w5 = Tk()
        w5.title(clicked5.get())
        w5.geometry('600x280+300+200')
        w5.resizable(0, 0)
        frame4 = Frame(w5)
        frame4.pack(side=TOP, expand=False)
        Label(frame4, text=clicked5.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame4, text='f(x) = ', font=(
            "Times", "12")).grid(row=1, column=0)
        D1 = Entry(frame4, width=50)
        D1.grid(row=1, column=1, columnspan=2)
        Label(frame4, text='Inital guess x0 ', font=(
            "Times", "12")).grid(row=2, column=0)
        D2 = Entry(frame4, width=50)
        D2.grid(row=2, column=1, columnspan=2)
        Label(frame4, text='Inital guess x1 ', font=(
            "Times", "12")).grid(row=3, column=0)
        D3 = Entry(frame4, width=50)
        D3.grid(row=3, column=1, columnspan=2)
        Label(frame4, text='Inital guess x2 ', font=(
            "Times", "12")).grid(row=4, column=0)
        D4 = Entry(frame4, width=50)
        D4.grid(row=4, column=1, columnspan=2)
        Button(frame4, text='calculate max', width=12,
               command=parabolic_interpolation).grid(row=5, column=1)
    if clicked5.get() == "Newton's Method":
        w5 = Tk()
        w5.title(clicked5.get())
        w5.geometry('600x250+300+200')
        w5.resizable(0, 0)
        frame4 = Frame(w5)
        frame4.pack(side=TOP, expand=False)
        Label(frame4, text=clicked5.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame4, text='f(x) = ', font=(
            "Times", "12")).grid(row=1, column=0)
        D1 = Entry(frame4, width=50)
        D1.grid(row=1, column=1, columnspan=2)
        Label(frame4, text='Inital guess x0 ', font=(
            "Times", "12")).grid(row=2, column=0)
        D2 = Entry(frame4, width=50)
        D2.grid(row=2, column=1, columnspan=2)
        Button(frame4, text='calculate max', width=12,
               command=newt_optimization).grid(row=3, column=1)
    if clicked5.get() == 'Optimal Steepest Ascent Method':
        w5 = Tk()
        w5.title(clicked5.get())
        w5.geometry('600x260+300+200')
        w5.resizable(0, 0)
        frame4 = Frame(w5)
        frame4.pack(side=TOP, expand=False)
        Label(frame4, text=clicked5.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame4, text='f(x,y) = ', font=(
            "Times", "12")).grid(row=1, column=0)
        D1 = Entry(frame4, width=50)
        D1.grid(row=1, column=1, columnspan=2)
        Label(frame4, text='Inital guess x ', font=(
            "Times", "12")).grid(row=2, column=0)
        D2 = Entry(frame4, width=50)
        D2.grid(row=2, column=1, columnspan=2)
        Label(frame4, text='Inital guess y ', font=(
            "Times", "12")).grid(row=3, column=0)
        D3 = Entry(frame4, width=50)
        D3.grid(row=3, column=1, columnspan=2)
        Button(frame4, text='calculate max', width=12,
               command=optimal).grid(row=4, column=1)


def optimal():
    unconstrained.steep_ascent(D1.get(), D2.get(), D3.get(), frame4)


def newt_optimization():
    if D1.get() == '' or D2.get() == '':
        messagebox.showerror('Error', 'An Input Box is Empty')
        root.lower()
        pass
    else:
        try:
            unconstrained.newt_optim(D1.get(), D2.get(), frame4)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def maxim_gold():
    if D1.get() == '' or D2.get() == '' or D3.get == '':
        messagebox.showerror('Error', 'An Input Box is Empty')
        root.lower()
        pass
    else:
        try:
            unconstrained.golden_max(D1.get(), D2.get(), D3.get(), frame4)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def minim_gold():
    if D1.get() == '' or D2.get() == '' or D3.get == '':
        messagebox.showerror('Error', 'An Input Box is Empty')
        root.lower()
        pass
    else:
        try:
            unconstrained.golden_min(D1.get(), D2.get(), D3.get(), frame4)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass
    #   2*sin(x)-(x**2/10)


def parabolic_interpolation():
    if D1.get() == '' or D2.get() == '' or D3.get == '' or D4.get() == '':
        messagebox.showerror('Error', 'An Input Box is Empty')
        root.lower()
        pass
    else:
        try:
            unconstrained.parab_interp(
                D1.get(), D2.get(), D3.get(), D4.get(), frame4)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def b1_press():
    global w2, frame2, B1, B2, B3, B4, B5
    if clicked2.get() == "Euler's Method":
        w2 = Tk()
        w2.title(clicked2.get())
        w2.geometry('600x250+300+200')
        w2.resizable(0, 0)
        frame2 = Frame(w2)
        frame2.pack(side=TOP, expand=False)
        Label(frame2, text=clicked2.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame2, text='dy/dx =', font=("Times", "12")).grid(row=1, column=0)
        B1 = Entry(frame2, width=50)
        B1.grid(row=1, column=1, columnspan=2)
        Label(frame2, text='Inital x-val',
              font=("Times", "12")).grid(row=2, column=0)
        B2 = Entry(frame2, width=50)
        B2.grid(row=2, column=1, columnspan=2)
        Label(frame2, text='Inital y-val',
              font=("Times", "12")).grid(row=3, column=0)
        B3 = Entry(frame2, width=50)
        B3.grid(row=3, column=1, columnspan=2)
        Label(frame2, text='Step size', font=(
            "Times", "12")).grid(row=4, column=0)
        B4 = Entry(frame2, width=50)
        B4.grid(row=4, column=1, columnspan=2)
        Label(frame2, text='Evaluate at x =', font=(
            "Times", "12")).grid(row=5, column=0)
        B5 = Entry(frame2, width=50)
        B5.grid(row=5, column=1, columnspan=2)
        button = Button(frame2, text='calculate', width=10,
                        command=eu_method).grid(row=6, column=0)
    if clicked2.get() == "Heun's Method":
        w2 = Tk()
        w2.title(clicked2.get())
        w2.geometry('600x250+300+200')
        w2.resizable(0, 0)
        frame2 = Frame(w2)
        frame2.pack(side=TOP, expand=False)
        Label(frame2, text=clicked2.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame2, text='dy/dx =', font=("Times", "12")).grid(row=1, column=0)
        B1 = Entry(frame2, width=50)
        B1.grid(row=1, column=1, columnspan=2)
        Label(frame2, text='Inital x-val',
              font=("Times", "12")).grid(row=2, column=0)
        B2 = Entry(frame2, width=50)
        B2.grid(row=2, column=1, columnspan=2)
        Label(frame2, text='Inital y-val',
              font=("Times", "12")).grid(row=3, column=0)
        B3 = Entry(frame2, width=50)
        B3.grid(row=3, column=1, columnspan=2)
        Label(frame2, text='Step size', font=(
            "Times", "12")).grid(row=4, column=0)
        B4 = Entry(frame2, width=50)
        B4.grid(row=4, column=1, columnspan=2)
        Label(frame2, text='Evaluate at x =', font=(
            "Times", "12")).grid(row=5, column=0)
        B5 = Entry(frame2, width=50)
        B5.grid(row=5, column=1, columnspan=2)
        button = Button(frame2, text='calculate', width=10,
                        command=heun_method).grid(row=6, column=0)
    if clicked2.get() == 'Midpoint Method':
        w2 = Tk()
        w2.title(clicked2.get())
        w2.geometry('600x250+300+200')
        w2.resizable(0, 0)
        frame2 = Frame(w2)
        frame2.pack(side=TOP, expand=False)
        Label(frame2, text=clicked2.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame2, text='dy/dx =', font=("Times", "12")).grid(row=1, column=0)
        B1 = Entry(frame2, width=50)
        B1.grid(row=1, column=1, columnspan=2)
        Label(frame2, text='Inital x-val',
              font=("Times", "12")).grid(row=2, column=0)
        B2 = Entry(frame2, width=50)
        B2.grid(row=2, column=1, columnspan=2)
        Label(frame2, text='Inital y-val',
              font=("Times", "12")).grid(row=3, column=0)
        B3 = Entry(frame2, width=50)
        B3.grid(row=3, column=1, columnspan=2)
        Label(frame2, text='Step size', font=(
            "Times", "12")).grid(row=4, column=0)
        B4 = Entry(frame2, width=50)
        B4.grid(row=4, column=1, columnspan=2)
        Label(frame2, text='Evaluate at x =', font=(
            "Times", "12")).grid(row=5, column=0)
        B5 = Entry(frame2, width=50)
        B5.grid(row=5, column=1, columnspan=2)
        button = Button(frame2, text='calculate', width=10,
                        command=mp_method).grid(row=6, column=0)
    if clicked2.get() == '4th Order Runge-Kutta':
        w2 = Tk()
        w2.title(clicked2.get())
        w2.geometry('600x250+300+200')
        w2.resizable(0, 0)
        frame2 = Frame(w2)
        frame2.pack(side=TOP, expand=False)
        Label(frame2, text=clicked2.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame2, text='dy/dx =', font=("Times", "12")).grid(row=1, column=0)
        B1 = Entry(frame2, width=50)
        B1.grid(row=1, column=1, columnspan=2)
        Label(frame2, text='Inital x-val',
              font=("Times", "12")).grid(row=2, column=0)
        B2 = Entry(frame2, width=50)
        B2.grid(row=2, column=1, columnspan=2)
        Label(frame2, text='Inital y-val',
              font=("Times", "12")).grid(row=3, column=0)
        B3 = Entry(frame2, width=50)
        B3.grid(row=3, column=1, columnspan=2)
        Label(frame2, text='Step size', font=(
            "Times", "12")).grid(row=4, column=0)
        B4 = Entry(frame2, width=50)
        B4.grid(row=4, column=1, columnspan=2)
        Label(frame2, text='Evaluate at x =', font=(
            "Times", "12")).grid(row=5, column=0)
        B5 = Entry(frame2, width=50)
        B5.grid(row=5, column=1, columnspan=2)
        button = Button(frame2, text='calculate', width=10,
                        command=rk_method).grid(row=6, column=0)
    if clicked2.get() == '5th Order Runge-Kutta':
        w2 = Tk()
        w2.title(clicked2.get())
        w2.geometry('600x250+300+200')
        w2.resizable(0, 0)
        frame2 = Frame(w2)
        frame2.pack(side=TOP, expand=False)
        Label(frame2, text=clicked2.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame2, text='dy/dx =', font=("Times", "12")).grid(row=1, column=0)
        B1 = Entry(frame2, width=50)
        B1.grid(row=1, column=1, columnspan=2)
        Label(frame2, text='Inital x-val',
              font=("Times", "12")).grid(row=2, column=0)
        B2 = Entry(frame2, width=50)
        B2.grid(row=2, column=1, columnspan=2)
        Label(frame2, text='Inital y-val',
              font=("Times", "12")).grid(row=3, column=0)
        B3 = Entry(frame2, width=50)
        B3.grid(row=3, column=1, columnspan=2)
        Label(frame2, text='Step size', font=(
            "Times", "12")).grid(row=4, column=0)
        B4 = Entry(frame2, width=50)
        B4.grid(row=4, column=1, columnspan=2)
        Label(frame2, text='Evaluate at x =', font=(
            "Times", "12")).grid(row=5, column=0)
        B5 = Entry(frame2, width=50)
        B5.grid(row=5, column=1, columnspan=2)
        button = Button(frame2, text='calculate', width=10,
                        command=fifth_rk).grid(row=6, column=0)


def fifth_rk():
    try:
        differentiate.fif_rkutta(
            B1.get(), B5.get(), B4.get(), B2.get(), B3.get(), frame2, w2)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass


def rk_method():
    try:
        differentiate.runge_kutta(
            B1.get(), B5.get(), B4.get(), B2.get(), B3.get(), frame2, w2)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass


def mp_method():
    try:
        differentiate.midpoint(B1.get(), B5.get(), B4.get(),
                               B2.get(), B3.get(), frame2, w2)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass


def heun_method():
    try:
        differentiate.heun(B1.get(), B5.get(), B4.get(),
                           B2.get(), B3.get(), frame2, w2)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass


def eu_method():
    try:
        differentiate.euler(B1.get(), B5.get(), B4.get(),
                            B2.get(), B3.get(), frame2, w2)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass


def b2_press():
    global A1, A2, A3, A4, frame3, w3
    if clicked3.get() == 'Single Application Trapezoidal Rule':
        w3 = Tk()
        w3.title(clicked3.get())
        w3.geometry('600x200+300+200')
        w3.resizable(0, 0)
        frame3 = Frame(w3)
        frame3.pack(side=TOP, expand=False)
        Label(frame3, text=clicked3.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame3, text='Integrate f(x) =', font=(
            "Times", "12")).grid(row=1, column=0)
        A1 = Entry(frame3, width=50)
        A1.grid(row=1, column=1, columnspan=2)
        Label(frame3, text='From a = ', font=(
            "Times", "12")).grid(row=2, column=0)
        A2 = Entry(frame3, width=50)
        A2.grid(row=2, column=1, columnspan=2)
        Label(frame3, text='To b = ', font=(
            "Times", "12")).grid(row=3, column=0)
        A3 = Entry(frame3, width=50)
        A3.grid(row=3, column=1, columnspan=2)
        button = Button(frame3, text='calculate', width=10,
                        command=sap_trap).grid(row=5, column=0)
    if clicked3.get() == 'Multiple Application Trapezoidal Rule':
        w3 = Tk()
        w3.title(clicked3.get())
        w3.geometry('600x220+300+200')
        w3.resizable(0, 0)
        frame3 = Frame(w3)
        frame3.pack(side=TOP, expand=False)
        Label(frame3, text=clicked3.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame3, text='Integrate f(x) =', font=(
            "Times", "12")).grid(row=1, column=0)
        A1 = Entry(frame3, width=50)
        A1.grid(row=1, column=1, columnspan=2)
        Label(frame3, text='From a = ', font=(
            "Times", "12")).grid(row=2, column=0)
        A2 = Entry(frame3, width=50)
        A2.grid(row=2, column=1, columnspan=2)
        Label(frame3, text='To b = ', font=(
            "Times", "12")).grid(row=3, column=0)
        A3 = Entry(frame3, width=50)
        A3.grid(row=3, column=1, columnspan=2)
        Label(frame3, text='Segements n = ', font=(
            "Times", "12")).grid(row=4, column=0)
        A4 = Entry(frame3, width=50)
        A4.grid(row=4, column=1, columnspan=2)
        button = Button(frame3, text='calculate', width=10,
                        command=map_trap).grid(row=5, column=0)
    if clicked3.get() == "Single Application Simpson's 1/3 Rule":
        w3 = Tk()
        w3.title(clicked3.get())
        w3.geometry('600x200+300+200')
        w3.resizable(0, 0)
        frame3 = Frame(w3)
        frame3.pack(side=TOP, expand=False)
        Label(frame3, text=clicked3.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame3, text='Integrate f(x) =', font=(
            "Times", "12")).grid(row=1, column=0)
        A1 = Entry(frame3, width=50)
        A1.grid(row=1, column=1, columnspan=2)
        Label(frame3, text='From a = ', font=(
            "Times", "12")).grid(row=2, column=0)
        A2 = Entry(frame3, width=50)
        A2.grid(row=2, column=1, columnspan=2)
        Label(frame3, text='To b = ', font=(
            "Times", "12")).grid(row=3, column=0)
        A3 = Entry(frame3, width=50)
        A3.grid(row=3, column=1, columnspan=2)
        button = Button(frame3, text='calculate', width=10,
                        command=sap_one_third).grid(row=5, column=0)
    if clicked3.get() == "Multiple Application Simpson's 1/3 Rule":
        w3 = Tk()
        w3.title(clicked3.get())
        w3.geometry('600x220+300+200')
        w3.resizable(0, 0)
        frame3 = Frame(w3)
        frame3.pack(side=TOP, expand=False)
        Label(frame3, text=clicked3.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame3, text='Integrate f(x) =', font=(
            "Times", "12")).grid(row=1, column=0)
        A1 = Entry(frame3, width=50)
        A1.grid(row=1, column=1, columnspan=2)
        Label(frame3, text='From a = ', font=(
            "Times", "12")).grid(row=2, column=0)
        A2 = Entry(frame3, width=50)
        A2.grid(row=2, column=1, columnspan=2)
        Label(frame3, text='To b = ', font=(
            "Times", "12")).grid(row=3, column=0)
        A3 = Entry(frame3, width=50)
        A3.grid(row=3, column=1, columnspan=2)
        Label(frame3, text='Segements n = ', font=(
            "Times", "12")).grid(row=4, column=0)
        A4 = Entry(frame3, width=50)
        A4.grid(row=4, column=1, columnspan=2)
        button = Button(frame3, text='calculate', width=10,
                        command=map_one_third).grid(row=5, column=0)
    if clicked3.get() == "Simpson's 3/8 Rule":
        w3 = Tk()
        w3.title(clicked3.get())
        w3.geometry('600x200+300+200')
        w3.resizable(0, 0)
        frame3 = Frame(w3)
        frame3.pack(side=TOP, expand=False)
        Label(frame3, text=clicked3.get(), font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame3, text='Integrate f(x) =', font=(
            "Times", "12")).grid(row=1, column=0)
        A1 = Entry(frame3, width=50)
        A1.grid(row=1, column=1, columnspan=2)
        Label(frame3, text='From a = ', font=(
            "Times", "12")).grid(row=2, column=0)
        A2 = Entry(frame3, width=50)
        A2.grid(row=2, column=1, columnspan=2)
        Label(frame3, text='To b = ', font=(
            "Times", "12")).grid(row=3, column=0)
        A3 = Entry(frame3, width=50)
        A3.grid(row=3, column=1, columnspan=2)
        button = Button(frame3, text='calculate', width=10,
                        command=three_eighth).grid(row=5, column=0)
    if clicked3.get() == 'Trapezoidal Rule with Unequal Segments':
        w3 = Toplevel()
        w3.title(clicked3.get())
        w3.geometry('600x420+300+100')
        w3.resizable(0, 0)
        Label(w3, text='Select an excel file that contains a column of x values and a column of f(x) values', font=(
            "Times", "12")).pack()
        Label(w3, text='The data can be placed anywhere in the excel sheet,', font=(
            "Times", "12")).pack()
        Label(w3, text='but it must be grouped together as in the following example:', font=(
            "Times", "12")).pack()
        photo = ImageTk.PhotoImage(Image.open(
            r'c:\Users\ahalw\demo.jpg').resize((400, 250), Image.ANTIALIAS))
        lab = Label(w3, image=photo)
        lab.image = photo
        lab.pack()
        button = Button(w3, text='Select File',
                        width=14, command=unequal).pack()


def unequal():
    w3.filename = filedialog.askopenfilename(
        initialdir=os.path.dirname(__file__), title="Select A File", filetypes=(("excel files", "*.xlsx"), ("all files", "*.")))
    if w3.filename == '':
        root.lower()
        pass
    else:
        wb = load_workbook(filename=w3.filename)
        ws = wb['Sheet1']
        x = []
        fx = []
        stop = false
        for column in ws.iter_cols(1, ws.max_column):  # iterate column cell
            for row in range(len(column)):
                if column[row].value == 'x':    # check for your column
                    for data in column[row+1:]:    # iterate your column
                        x.append(data.value)
                if column[row].value == 'f(x)':    # check for your column
                    print(column[row].value)
                    stop = true
                    for data in column[row+1:]:    # iterate your column
                        fx.append(data.value)
            if stop == true:
                break
        if len(x) != len(fx):
            messagebox.showerror(
                'Error', 'The number of entries in x and f(x) must be equal')
            root.lower()
            pass
        try:
            integrate.uneq_segments(x, fx, w3)
            root.lower()
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def three_eighth():
    try:
        integrate.threeEighth(A1.get(), A2.get(), A3.get(), frame3, w3)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass


def sap_one_third():
    try:
        integrate.sing_onethird(A1.get(), A2.get(), A3.get(), frame3, w3)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass


def map_one_third():
    if int(A4.get()) < 4:
        messagebox.showerror(
            title='Error', message="Segments n can't be less that 4")
        root.lower()
    else:
        try:
            integrate.mult_onethird(
                A1.get(), A2.get(), A3.get(), A4.get(), frame3, w3)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def sap_trap():
    try:
        integrate.sing_trap(A1.get(), A2.get(), A3.get(), frame3, w3)
    except BaseException:
        messagebox.showerror('Error', 'Invalid Input Syntax')
        root.lower()
        pass

    # 0.2+25*x-200*x^(2)+675*x^(3)-900*x^(4)+400*x^(5)


def map_trap():
    if int(A4.get()) < 2:
        messagebox.showerror(
            title='Error', message="Segments n can't be less that 2")
        root.lower()
    else:
        try:
            integrate.mult_trap(A1.get(), A2.get(),
                                A3.get(), A4.get(), frame3, w3)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def b3_press():
    global C1, C2, C3, w4
    if clicked4.get() == 'Linear Regression':
        w4 = Toplevel()
        w4.title(clicked4.get())
        w4.geometry('600x500+300+100')
        w4.resizable(0, 0)
        Label(w4, text='Select an excel file that contains a column of x values and a column of y values', font=(
            "Times", "12")).pack()
        Label(w4, text='The data can be placed anywhere in the excel sheet,', font=(
            "Times", "12")).pack()
        Label(w4, text='but it must be grouped together as in the following example:', font=(
            "Times", "12")).pack()
        directory_path = os.path.dirname(__file__)
        file_path = os.path.join(
            directory_path, 'demo2.jpg')
        photo = ImageTk.PhotoImage(Image.open(
            file_path).resize((400, 250), Image.ANTIALIAS))
        lab = Label(w4, image=photo)
        lab.image = photo
        lab.pack()
        button = Button(w4, text='Select File', width=14,
                        command=linear_reg).pack()
    if clicked4.get() == 'Polynomial Regression':
        w4 = Toplevel()
        w4.title(clicked4.get())
        w4.geometry('700x550+300+100')
        w4.resizable(0, 0)
        Label(w4, text='Select an excel file that contains a column of x values and a column of y values', font=(
            "Times", "12")).pack()
        Label(w4, text='The data can be placed anywhere in the excel sheet,', font=(
            "Times", "12")).pack()
        Label(w4, text='but it must be grouped together as in the following example:', font=(
            "Times", "12")).pack()
        directory_path = os.path.dirname(__file__)
        file_path = os.path.join(
            directory_path, 'demo2.jpg')
        photo = ImageTk.PhotoImage(Image.open(
            file_path).resize((400, 250), Image.ANTIALIAS))
        lab = Label(w4, image=photo)
        lab.image = photo
        lab.pack()
        Label(w4, text='Enter the polynomial order below as an integer:').pack()
        C1 = Entry(w4, width=20)
        C1.pack()
        button = Button(w4, text='Select File', width=14,
                        command=polynomial_reg).pack()
    if clicked4.get() == 'Multiple Linear Regression':
        w4 = Toplevel()
        w4.title(clicked4.get())
        w4.geometry('700x550+300+100')
        w4.resizable(0, 0)
        Label(w4, text='Select an excel file that contains a column of x values and a column of y values', font=(
            "Times", "12")).pack()
        Label(w4, text='The data can be placed anywhere in the excel sheet,', font=(
            "Times", "12")).pack()
        Label(w4, text='but it must be grouped together as in the following example:', font=(
            "Times", "12")).pack()
        directory_path = os.path.dirname(__file__)
        file_path = os.path.join(
            directory_path, 'demo3.jpg')
        photo = ImageTk.PhotoImage(Image.open(
            file_path).resize((400, 250), Image.ANTIALIAS))
        lab = Label(w4, image=photo)
        lab.image = photo
        lab.pack()
        Label(w4, text='Number of Independent Variables:').pack()
        C1 = Entry(w4, width=20)
        C1.pack()
        button = Button(w4, text='Select File', width=14,
                        command=multiple_reg).pack()
    if clicked4.get() == 'Non-Linear Regression':
        w4 = Toplevel()
        w4.title(clicked4.get())
        w4.geometry('700x520+300+70')
        w4.resizable(0, 0)
        Label(w4, text='Select an excel file that contains a column of x values and a column of y values', font=(
            "Times", "12")).pack()
        Label(w4, text='The data can be placed anywhere in the excel sheet,', font=(
            "Times", "12")).pack()
        Label(w4, text='but it must be grouped together as in the following example:', font=(
            "Times", "12")).pack()
        directory_path = os.path.dirname(__file__)
        file_path = os.path.join(
            directory_path, 'demo.jpg')
        photo = ImageTk.PhotoImage(Image.open(
            file_path).resize((400, 250), Image.ANTIALIAS))
        lab = Label(w4, image=photo)
        lab.image = photo
        lab.pack()
        Label(w4, text='Parameters seperated by Witespace (e.g. a0 a1 a2):').pack()
        C1 = Entry(w4, width=30)
        C1.pack()
        Label(w4, text='Corresponding Initial Guesses seperated by Witespace (e.g. 1 2 3):').pack()
        C2 = Entry(w4, width=30)
        C2.pack()
        Label(w4, text='Function to fit the data to:').pack()
        C3 = Entry(w4, width=40)
        C3.pack()
        button = Button(w4, text='Select File',
                        width=14, command=non_reg).pack()


def non_reg():
    w4.filename = filedialog.askopenfilename(
        initialdir=os.path.dirname(__file__), title="Select A File", filetypes=(("excel files", "*.xlsx"), ("all files", "*.")))
    if w4.filename == '':
        root.lower()
        pass
    else:
        wb = load_workbook(filename=w4.filename)
        ws = wb['Sheet1']
        x = []
        fx = []
        stop = false
        for column in ws.iter_cols(1, ws.max_column):  # iterate column cell
            for row in range(len(column)):
                if column[row].value == 'x':    # check for your column
                    for data in column[row+1:]:    # iterate your column
                        x.append(data.value)
                if column[row].value == 'f(x)':    # check for your column
                    print(column[row].value)
                    stop = true
                    for data in column[row+1:]:    # iterate your column
                        fx.append(data.value)
            if stop == true:
                break
        if len(x) != len(fx):
            messagebox.showerror(
                'Error', 'The number of entries in x and f(x) must be equal')
            root.lower()
            pass
        try:
            curve_fitting.nonlin_reg(x, fx, C3.get(), C1.get(), C2.get())
            root.lower()
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def multiple_reg():
    w4.filename = filedialog.askopenfilename(
        initialdir=os.path.dirname(__file__), title="Select A File", filetypes=(("excel files", "*.xlsx"), ("all files", "*.")))
    if w4.filename == '':
        root.lower()
        pass
    else:
        wb = load_workbook(filename=w4.filename)
        ws = wb['Sheet1']
        nvars = 0
        x = []
        y = []
        stop = false
        for column in ws.iter_cols(1, ws.max_column):  # iterate column cell
            holder = []
            for row in range(len(column)):
                if 'x' in str(column[row].value):    # check for your column
                    for data in column[row+1:]:    # iterate your column
                        holder.append(data.value)
                    x.append(holder)
                    nvars += 1
                if column[row].value == 'y':    # check for your column
                    print(column[row].value)
                    stop = true
                    for data in column[row+1:]:    # iterate your column
                        y.append(data.value)
            if stop == true:
                break
        if len(x[1]) != len(y):
            messagebox.showerror(
                'Error', 'The number of entries in x and y must be equal')
            root.lower()
            pass
        try:
            curve_fitting.mult_reg(x, y, nvars, w4)
            root.lower()
        except IndexError:
            messagebox.showerror(
                'Error', "Number of Independent Variables doesn't match datasheet")
            root.lower()
            pass
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def polynomial_reg():
    w4.filename = filedialog.askopenfilename(
        initialdir=os.path.dirname(__file__), title="Select A File", filetypes=(("excel files", "*.xlsx"), ("all files", "*.")))
    if w4.filename == '':
        root.lower()
        pass
    else:
        wb = load_workbook(filename=w4.filename)
        ws = wb['Sheet1']
        x = []
        y = []
        stop = false
        for column in ws.iter_cols(1, ws.max_column):  # iterate column cell
            for row in range(len(column)):
                if column[row].value == 'x':    # check for your column
                    for data in column[row+1:]:    # iterate your column
                        x.append(data.value)
                if column[row].value == 'y':    # check for your column
                    print(column[row].value)
                    stop = true
                    for data in column[row+1:]:    # iterate your column
                        y.append(data.value)
            if stop == true:
                break
        if len(x) != len(y):
            messagebox.showerror(
                'Error', 'The number of entries in x and y must be equal')
            root.lower()
            pass
        try:
            curve_fitting.pol_reg(x, y, C1.get(), w4)
            root.lower()
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def linear_reg():
    w4.filename = filedialog.askopenfilename(
        initialdir=os.path.dirname(__file__), title="Select A File", filetypes=(("excel files", "*.xlsx"), ("all files", "*.")))
    if w4.filename == '':
        root.lower()
        pass
    else:
        wb = load_workbook(filename=w4.filename)
        ws = wb['Sheet1']
        x = []
        y = []
        stop = false
        for column in ws.iter_cols(1, ws.max_column):  # iterate column cell
            for row in range(len(column)):
                if column[row].value == 'x':    # check for your column
                    for data in column[row+1:]:    # iterate your column
                        x.append(data.value)
                if column[row].value == 'y':    # check for your column
                    print(column[row].value)
                    stop = true
                    for data in column[row+1:]:    # iterate your column
                        y.append(data.value)
            if stop == true:
                break
        if len(x) != len(y):
            messagebox.showerror(
                'Error', 'The number of entries in x and f(x) must be equal')
            root.lower()
            pass
        try:
            curve_fitting.lin_reg(x, y, w4)
            root.lower()
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def mod_newt_raphs():
    global pressed3
    pressed3 += 1
    if E1.get() == '' and E2.get() == '' and E3.get() == '' and E4.get() == '':
        messagebox.showerror(title='Error', message='Empty Input Box!')
        root.lower()
    else:
        try:
            execute(pressed3).mod_newt(E1.get(), E2.get(), E3.get(), frame)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def newt_raphs():
    global pressed
    pressed += 1
    if E1.get() == '' and E2.get() == '' and E3.get() == '' and E4.get() == '':
        messagebox.showerror(title='Error', message='Empty Input Box!')
        root.lower()
    else:
        try:
            execute(pressed).newt(E1.get(), E2.get(), E3.get(), frame)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass

    return


def mod_secant():
    global pressed1
    pressed1 += 1
    if E1.get() == '' and E2.get() == '' and E3.get() == '' and E4.get() == '':
        messagebox.showerror(title='Error', message='Empty Input Box!')
        root.lower()
    else:
        try:
            execute(pressed1).mod_finish(
                E1.get(), E2.get(), E3.get(), E4.get(), frame)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


def secant():
    global pressed2
    pressed2 += 1
    if E1.get() == '' and E2.get() == '' and E3.get() == '' and E4.get() == '':
        messagebox.showerror(title='Error', message='Empty Input Box!')
        root.lower()
    else:
        try:
            execute(pressed2).finish(
                E1.get(), E2.get(), E3.get(), E4.get(), frame)
        except BaseException:
            messagebox.showerror('Error', 'Invalid Input Syntax')
            root.lower()
            pass


class system:
    def __init__(self, c, entries, w1, frame):
        self.c = c
        self.entries = entries
        self.w1 = w1
        self.frame = frame
        self.inp = []

    def butt1(self):
        self.w1.geometry = ''
        con = Entry(self.frame, width=50)
        con.grid(row=self.c+1, column=1, columnspan=2)
        self.entries.append(con)
        self.c += 1

    def _delete_window(self):
        self.c = 1
        try:
            self.w1.destroy()
        except:
            pass

    def make_inp(self):
        for entry in self.entries:
            self.inp.append(entry.get())
        if self.inp[0] == '':
            messagebox.showerror('Error', 'Must Enter An Equation')
            w1.lift()
        else:
            for i in range(len(self.inp)):
                if i != 0 and self.inp[i] == '':
                    self.inp.remove('')
            # declare variables
            frame.destroy()
            nframe = Frame(w1)
            nframe.pack(side=TOP, expand=False)
            Label(nframe, text='Seperate each Entry by a Whitespace',
                  font=("Times", "14", "bold")).grid(row=0, column=1)
            Label(nframe, text='Variables', font=(
                "Times", "12")).grid(row=1, column=0)
            E1 = Entry(nframe, width=50)
            E1.grid(row=1, column=1, columnspan=2)
            Label(nframe, text='Inital Guesses', font=(
                "Times", "12")).grid(row=2, column=0)
            E2 = Entry(nframe, width=50)
            E2.grid(row=2, column=1, columnspan=2)
            Label(nframe, text='Appx. % Relative Error',
                  font=("Times", "12")).grid(row=3, column=0)
            E3 = Entry(nframe, width=50)
            E3.grid(row=3, column=1, columnspan=2)
            Button(nframe, text='Solve', width=10, command=lambda: system.solve(
                self, E1.get(), E2.get(), E3.get())).grid(row=4, column=0)

    def solve(self, v, IG, err):
        perror = float(err)
        ig = IG.split()
        for i in range(len(ig)):
            ig[i] = float(ig[i])
        x = v.split()
        if len(ig) != len(self.inp) or len(x) != len(self.inp):
            messagebox.showerror(
                'Error', 'Number of Variables/Inital guesses\nMust Match Number of Equations')
            w1.lift()
        else:
            if len(ig) != len(x):
                messagebox.showerror(
                    'Error', 'Number of Variables and Inital Guesses must be Equal')
                w1.lift()
        for i in range(len(x)):
            x[i] = sym.symbols(x[i])

        while True:
            try:
                a = execute(0).sys_eq(self.inp, x, ig)
            except TypeError:
                messagebox.showerror('Error', 'Invalid Input')
                root.lower()
                pass
            error = a[0]
            ig = a[1]
            if error <= perror:
                w2 = Tk()
                w2.title('Solution')
                w2.geometry('200x200+300+400')
                w2.resizable(0, 0)
                w2.geometry('')
                Label(w2, text='Solution', font=("Times", "10")
                      ).pack()  # grid(row=5,column=0)
                r = 6
                for i in range(len(x)):
                    # .grid(row=r,column=0)
                    Label(w2, text=str(x[i])+': '+str(ig[i]),
                          font=("Times", "10")).pack()
                    r += 1
                Label(w2, text='Error: '+str(error)+'%',
                      font=("Times", "10")).pack()  # .grid(row=r,column=0)
                break


def b_press():
    global E1, E2, E3, E4, frame, w1
    pressed = 0
    if clicked.get() == 'Non-Linear System of Eqs':
        w1 = Tk()
        w1.title(clicked.get())
        w1.geometry('600x200+300+200')
        w1.resizable(0, 0)
        frame = Frame(w1)
        frame.pack(side=TOP, expand=False)
        Label(frame, text='System of Equations Calculator', font=(
            "Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame, text='Equation(s) ', font=(
            "Times", "12")).grid(row=1, column=0)
        E1 = Entry(frame, width=50)
        E1.grid(row=1, column=1, columnspan=2)
        entries.append(E1)
        button1 = Button(frame, text='Add Equation', width=14,
                         command=lambda: obj.butt1())
        button1.grid(row=2, column=0)
        Button(frame, text='Next Step', width=14,
               command=lambda: obj.make_inp()).grid(row=3, column=0)
        w1.protocol("WM_DELETE_WINDOW", lambda: obj._delete_window())
        obj = system(1, entries, w1, frame)
    if clicked.get() == 'Secant Method':
        w1 = Tk()
        w1.title(clicked.get())
        w1.geometry('550x200+300+200')
        w1.resizable(0, 0)
        frame = Frame(w1)
        frame.pack(side=TOP, expand=False)
        Label(frame, text=clicked.get()+' Calculator',
              font=("Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame, text='f(x) =', font=("Times", "12")).grid(row=1, column=0)
        E1 = Entry(frame, width=50)
        E1.grid(row=1, column=1, columnspan=2)
        l = Text(frame, width=18, height=1, borderwidth=0,
                 background=w1.cget("background"), font=("Times", "12"))
        l.tag_configure("subscript", offset=-4)
        l.insert("insert", "Initial guess x", "", "-1", "subscript", ":")
        l.configure(state="disabled")
        l.grid(row=2, column=0)
        E2 = Entry(frame, width=50)
        E2.grid(row=2, column=1, columnspan=2)
        h = Text(frame, width=17, height=1, borderwidth=0,
                 background=w1.cget("background"), font=("Times", "12"))
        h.tag_configure("subscript", offset=-4)
        h.insert("insert", "Initial guess x", "", "0", "subscript", ":")
        h.configure(state="disabled")
        h.grid(row=3, column=0)
        E3 = Entry(frame, width=50)
        E3.grid(row=3, column=1, columnspan=2)
        Label(frame, text='Relative Error [%]:', font=(
            "Times", "12")).grid(row=4, column=0)
        E4 = Entry(frame, width=50)
        E4.grid(row=4, column=1, columnspan=2)
        button = Button(frame, text='calculate', width=10,
                        command=secant).grid(row=5, column=1)
    if clicked.get() == 'Modified Secant':
        w1 = Tk()
        w1.title(clicked.get())
        w1.geometry('600x200+300+200')
        w1.resizable(0, 0)
        frame = Frame(w1)
        frame.pack(side=TOP, expand=False)
        Label(frame, text=clicked.get()+' Calculator',
              font=("Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame, text='f(x) =', font=("Times", "12")).grid(row=1, column=0)
        E1 = Entry(frame, width=50)
        E1.grid(row=1, column=1, columnspan=2)
        l = Text(frame, width=17, height=1, borderwidth=0,
                 background=w1.cget("background"), font=("Times", "12"))
        l.tag_configure("subscript", offset=-4)
        l.insert("insert", "Initial guess x", "", "0", "subscript", ":")
        l.configure(state="disabled")
        l.grid(row=2, column=0)
        E2 = Entry(frame, width=50)
        E2.grid(row=2, column=1, columnspan=2)
        Label(frame, text='Perturbation Fraction \u03b4:',
              font=("Times", "12")).grid(row=3, column=0)
        E3 = Entry(frame, width=50)
        E3.grid(row=3, column=1, columnspan=2)
        Label(frame, text='Relative Error [%]:', font=(
            "Times", "12")).grid(row=4, column=0)
        E4 = Entry(frame, width=50)
        E4.grid(row=4, column=1, columnspan=2)
        button = Button(frame, text='calculate', width=10,
                        command=mod_secant).grid(row=5, column=1)
    if clicked.get() == 'Newton-Raphson Method':
        w1 = Tk()
        w1.title(clicked.get())
        w1.geometry('600x200+300+200')
        w1.resizable(0, 0)
        frame = Frame(w1)
        frame.pack(side=TOP, expand=False)
        Label(frame, text=clicked.get()+' Calculator',
              font=("Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame, text='f(x) =', font=("Times", "12")).grid(row=1, column=0)
        E1 = Entry(frame, width=50)
        E1.grid(row=1, column=1, columnspan=2)
        l = Text(frame, width=17, height=1, borderwidth=0,
                 background=w1.cget("background"), font=("Times", "12"))
        l.tag_configure("subscript", offset=-4)
        l.insert("insert", "Initial guess x", "", "0", "subscript", ":")
        l.configure(state="disabled")
        l.grid(row=2, column=0)
        E2 = Entry(frame, width=50)
        E2.grid(row=2, column=1, columnspan=2)
        Label(frame, text='Relative Error [%]:', font=(
            "Times", "12")).grid(row=3, column=0)
        E3 = Entry(frame, width=50)
        E3.grid(row=3, column=1, columnspan=2)
        button = Button(frame, text='calculate', width=10,
                        command=newt_raphs).grid(row=4, column=1)
    if clicked.get() == 'Modified Newton-Raphson':
        w1 = Tk()
        w1.title(clicked.get())
        w1.geometry('600x200+300+200')
        w1.resizable(0, 0)
        frame = Frame(w1)
        frame.pack(side=TOP, expand=False)
        Label(frame, text=clicked.get()+' Calculator',
              font=("Times", "18", "bold italic")).grid(row=0, column=1)
        Label(frame, text='f(x) =', font=("Times", "12")).grid(row=1, column=0)
        E1 = Entry(frame, width=50)
        E1.grid(row=1, column=1, columnspan=2)
        l = Text(frame, width=17, height=1, borderwidth=0,
                 background=w1.cget("background"), font=("Times", "12"))
        l.tag_configure("subscript", offset=-4)
        l.insert("insert", "Initial guess x", "", "0", "subscript", ":")
        l.configure(state="disabled")
        l.grid(row=2, column=0)
        E2 = Entry(frame, width=50)
        E2.grid(row=2, column=1, columnspan=2)
        Label(frame, text='Relative Error [%]:', font=(
            "Times", "12")).grid(row=3, column=0)
        E3 = Entry(frame, width=50)
        E3.grid(row=3, column=1, columnspan=2)
        button = Button(frame, text='calculate', width=10,
                        command=mod_newt_raphs).grid(row=4, column=1)


def simp_method():
    simpo = simplex
    simpo.be()


clicked = StringVar()
clicked.set(options[0])
Label(root, text='').pack()
Label(root, text='Root Finding Methods').pack()
drop = OptionMenu(root, clicked, *options)
drop.pack()
Button(root, text='Launch', command=b_press).pack()

clicked2 = StringVar()
clicked2.set(options2[0])
Label(root, text='Numerical Differentiation').pack()
drop2 = OptionMenu(root, clicked2, *options2)
drop2.pack()
Button(root, text='Launch', command=b1_press).pack()

clicked3 = StringVar()
clicked3.set(options3[0])
Label(root, text='Numerical Integration').pack()
drop3 = OptionMenu(root, clicked3, *options3)
drop3.pack()
Button(root, text='Launch', command=b2_press).pack()

clicked4 = StringVar()
clicked4.set(options4[0])
Label(root, text='Curve Fitting Methods').pack()
drop4 = OptionMenu(root, clicked4, *options4)
drop4.pack()
Button(root, text='Launch', command=b3_press).pack()

clicked5 = StringVar()
clicked5.set(options5[0])
Label(root, text='Unconstrained Optimization').pack()
drop5 = OptionMenu(root, clicked5, *options5)
drop5.pack()
Button(root, text='Launch', command=b4_press).pack()

clicked6 = StringVar()
clicked6.set('Simplex Method')
Label(root, text='Constrained Optimization').pack()
drop6 = OptionMenu(root, clicked6, 'Simplex Method')
drop6.pack()
Button(root, text='Launch', command=simp_method).pack()

side_w = Tk()
side_w.attributes('-alpha', 0.8)  # transparent window
side_w.title('For Your Reference')
side_w.geometry('300x200+900+50')
side_w.resizable(0, 0)
Label(side_w, text='\n\nTrig Functions Representation\n' +
      'cosine: cos, sine: sin, tangent: tan\n\n' +
      'Inverse Trig Representation\n' +
      'arccosine: acos, arcsin: asin, arctangent: atan\n\n' +
      'Hyperbolic Functions\n' +
      'cosh, sinh, tanh\n\n' +
      'For Syntax Reference\n' +
      'Example of a function entry:\n' +
      '3*y^(2)*(2*x+e^(3*x))-log(3*y*x)\n\n').pack()
root.lift()
side_w.mainloop()
root.mainloop()
